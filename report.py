# =============================================================================
# report.py — Excel Report Generator
# Creates a formatted, color-coded Excel report + CSV for Power BI
# =============================================================================

import pandas as pd
import os
from datetime import datetime
import logging

logger = logging.getLogger(__name__)

# Grade → background color (hex)
GRADE_COLORS = {
    "A+": "1A7F5A",   # Deep green
    "A":  "27AE60",   # Green
    "B":  "F39C12",   # Amber
    "C":  "E67E22",   # Orange
    "D":  "C0392B",   # Red
}

SCORE_GRADIENT = [
    (85, "1A7F5A"),   # 85–100 → dark green
    (75, "27AE60"),   # 75–84  → green
    (60, "F39C12"),   # 60–74  → amber
    (45, "E67E22"),   # 45–59  → orange
    (0,  "C0392B"),   # 0–44   → red
]


def score_to_color(score: float) -> str:
    for threshold, color in SCORE_GRADIENT:
        if score >= threshold:
            return color
    return "C0392B"


def build_summary_dataframe(results: list[dict]) -> pd.DataFrame:
    """
    Flatten the scored results into a clean DataFrame for Excel / Power BI.
    """
    rows = []
    for r in results:
        tech = r.get("tech_indicators", {})
        fund = r.get("fundamentals", {})
        score = r.get("score", {})

        row = {
            # Identity
            "Symbol":           r["symbol"],
            "Company":          fund.get("company_name", r["symbol"]),
            "Sector":           fund.get("sector", "Unknown"),
            "Industry":         fund.get("industry", "Unknown"),

            # Score
            "Total Score":      score.get("total_score", 0),
            "Tech Score":       score.get("tech_score", 0),
            "Fund Score":       score.get("fund_score", 0),
            "Grade":            score.get("grade", "D"),
            "Flags":            score.get("flags", ""),

            # Price & Trend
            "Close (₹)":        tech.get("close"),
            "Day Change %":     tech.get("day_change_pct"),
            "EMA 20":           tech.get("ema20"),
            "EMA 50":           tech.get("ema50"),
            "EMA 200":          tech.get("ema200"),
            "EMA Stack ✓":     "✓" if tech.get("signal_ema_stack") else "✗",
            "Supertrend":       "Buy" if tech.get("signal_supertrend") else "Sell",
            "% from 52W High":  tech.get("pct_from_52w_high"),

            # Momentum
            "RSI (14)":         tech.get("rsi"),
            "MACD":             tech.get("macd"),
            "MACD Signal":      tech.get("macd_signal"),
            "MACD Histogram":   tech.get("macd_histogram"),
            "MACD Cross ✓":    "✓" if tech.get("signal_macd_cross") else "",
            "ADX":              tech.get("adx"),

            # Volume
            "Volume":           tech.get("volume"),
            "20D Avg Volume":   tech.get("vol_20d_avg"),
            "Volume Ratio":     tech.get("vol_ratio"),

            # Fundamentals
            "Market Cap (Cr)":  fund.get("market_cap_cr"),
            "P/E Ratio":        fund.get("pe_ratio"),
            "Forward P/E":      fund.get("forward_pe"),
            "ROE %":            fund.get("roe"),
            "Debt/Equity":      fund.get("debt_to_equity"),
            "Revenue Growth %": fund.get("revenue_growth"),
            "Earnings Growth %":fund.get("earnings_growth"),
            "Price/Book":       fund.get("price_to_book"),
            "Dividend Yield %": fund.get("dividend_yield"),
            "Current Ratio":    fund.get("current_ratio"),

            # Meta
            "Screener Link":    f"https://www.screener.in/company/{r['symbol']}/",
            "TradingView Link": f"https://www.tradingview.com/chart/?symbol=NSE:{r['symbol']}",
        }
        rows.append(row)

    df = pd.DataFrame(rows)
    df = df.sort_values("Total Score", ascending=False).reset_index(drop=True)
    return df


def generate_excel_report(results: list[dict], output_dir: str,
                           filename: str) -> str:
    """
    Generate formatted Excel report with multiple sheets.
    Returns the full output path.
    """
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)

    df_all = build_summary_dataframe(results)

    # Filtered views
    df_grade_a = df_all[df_all["Grade"].isin(["A+", "A"])].copy()
    df_grade_b = df_all[df_all["Grade"] == "B"].copy()
    df_watchlist = df_all[df_all["Total Score"] >= 60].copy()

    try:
        with pd.ExcelWriter(filepath, engine="openpyxl") as writer:
            # --- Sheet 1: Full Report ---
            df_all.to_excel(writer, sheet_name="📊 All Stocks", index=False)
            _format_sheet(writer, "📊 All Stocks", df_all)

            # --- Sheet 2: Grade A+ and A ---
            df_grade_a.to_excel(writer, sheet_name="🏆 Grade A", index=False)
            _format_sheet(writer, "🏆 Grade A", df_grade_a)

            # --- Sheet 3: Grade B Watch ---
            df_grade_b.to_excel(writer, sheet_name="👀 Grade B Watch", index=False)
            _format_sheet(writer, "👀 Grade B Watch", df_grade_b)

            # --- Sheet 4: Watchlist (score >= 60) ---
            df_watchlist.to_excel(writer, sheet_name="📋 Watchlist", index=False)
            _format_sheet(writer, "📋 Watchlist", df_watchlist)

            # --- Sheet 5: Summary Stats ---
            _write_summary_sheet(writer, df_all)

        logger.info(f"Excel report saved: {filepath}")
        return filepath

    except Exception as e:
        logger.error(f"Excel generation failed: {e}")
        raise


def _format_sheet(writer, sheet_name: str, df: pd.DataFrame):
    """Apply formatting to a worksheet."""
    try:
        from openpyxl.styles import (PatternFill, Font, Alignment,
                                     Border, Side)
        from openpyxl.utils import get_column_letter

        ws = writer.sheets[sheet_name]

        # Header row formatting
        header_fill = PatternFill("solid", fgColor="1C2833")
        header_font = Font(color="FFFFFF", bold=True, size=10)
        thin_border = Border(
            bottom=Side(style="thin", color="444444")
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center",
                                       wrap_text=True)
            cell.border = thin_border

        ws.row_dimensions[1].height = 30

        # Find column indices
        cols = {cell.value: cell.column for cell in ws[1]}

        # Color Grade column
        if "Grade" in cols:
            grade_col = cols["Grade"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[grade_col - 1]
                grade = cell.value
                if grade in GRADE_COLORS:
                    cell.fill = PatternFill("solid", fgColor=GRADE_COLORS[grade])
                    cell.font = Font(color="FFFFFF", bold=True)
                    cell.alignment = Alignment(horizontal="center")

        # Color Total Score column
        if "Total Score" in cols:
            score_col = cols["Total Score"]
            for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
                cell = row[score_col - 1]
                if cell.value is not None:
                    try:
                        color = score_to_color(float(cell.value))
                        cell.fill = PatternFill("solid", fgColor=color)
                        cell.font = Font(color="FFFFFF", bold=True)
                        cell.alignment = Alignment(horizontal="center")
                    except (TypeError, ValueError):
                        pass

        # Alternate row colors
        light_fill = PatternFill("solid", fgColor="F8F9FA")
        dark_fill  = PatternFill("solid", fgColor="EAECEE")
        for i, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), 1):
            for cell in row:
                if cell.column not in (cols.get("Grade"), cols.get("Total Score")):
                    cell.fill = light_fill if i % 2 == 0 else dark_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.font = Font(size=9)

        # Auto-fit columns
        for col_cells in ws.columns:
            max_len = 0
            col_letter = get_column_letter(col_cells[0].column)
            for cell in col_cells:
                try:
                    if cell.value:
                        max_len = max(max_len, len(str(cell.value)))
                except Exception:
                    pass
            ws.column_dimensions[col_letter].width = min(max(max_len + 2, 8), 30)

        # Freeze top row
        ws.freeze_panes = "A2"

        # Auto-filter
        ws.auto_filter.ref = ws.dimensions

    except Exception as e:
        logger.warning(f"Sheet formatting error on '{sheet_name}': {e}")


def _write_summary_sheet(writer, df_all: pd.DataFrame):
    """Write a summary stats sheet."""
    try:
        ws = writer.book.create_sheet("📈 Summary")
        run_date = datetime.now().strftime("%d-%b-%Y %H:%M IST")

        stats = [
            ["📊 SWING SCREENER — DAILY REPORT", ""],
            [f"Run Date: {run_date}", ""],
            ["", ""],
            ["OVERVIEW", "Count"],
            ["Total Stocks Analyzed", len(df_all)],
            ["Grade A+ Stocks",  len(df_all[df_all["Grade"] == "A+"])],
            ["Grade A Stocks",   len(df_all[df_all["Grade"] == "A"])],
            ["Grade B Stocks",   len(df_all[df_all["Grade"] == "B"])],
            ["Grade C Stocks",   len(df_all[df_all["Grade"] == "C"])],
            ["Grade D Stocks",   len(df_all[df_all["Grade"] == "D"])],
            ["", ""],
            ["SCORE STATS", "Value"],
            ["Average Score",    round(df_all["Total Score"].mean(), 1)],
            ["Highest Score",    df_all["Total Score"].max()],
            ["Lowest Score",     df_all["Total Score"].min()],
            ["", ""],
            ["TOP 5 STOCKS BY SCORE", "Score"],
        ]

        top5 = df_all.head(5)[["Symbol", "Total Score"]].values.tolist()
        stats.extend(top5)

        for row_data in stats:
            ws.append(row_data)

        writer.sheets["📈 Summary"] = ws

    except Exception as e:
        logger.warning(f"Summary sheet error: {e}")


def generate_csv(results: list[dict], output_dir: str, filename: str) -> str:
    """Generate Power BI-ready CSV."""
    os.makedirs(output_dir, exist_ok=True)
    filepath = os.path.join(output_dir, filename)
    df = build_summary_dataframe(results)

    # Add run date column for Power BI time-series tracking
    df.insert(0, "Run Date", datetime.now().strftime("%Y-%m-%d"))

    df.to_csv(filepath, index=False)
    logger.info(f"CSV saved: {filepath}")
    return filepath
